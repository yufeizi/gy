#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据传输管理器
功能：
1. 已搜索关键词和已点击商品的EXCEL文档传输到主目录
2. 每隔10分钟向主目录排队传输
3. EXCEL文档增量式保存
4. 已点击商品传输后主程序回传覆盖机制
"""

import os
import json
import time
import threading
# 🔥 修改：使用openpyxl替代pandas，减少依赖
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    try:
        import openpyxl
        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False
from datetime import datetime
from typing import Dict, List, Set, Optional
from pathlib import Path
import shutil
import platform

# 🔥 跨平台文件锁导入
try:
    import fcntl  # 文件锁（Linux/Mac）
    HAS_FCNTL = True
except ImportError:
    HAS_FCNTL = False

try:
    import msvcrt  # 文件锁（Windows）
    HAS_MSVCRT = True
except ImportError:
    HAS_MSVCRT = False


class DataTransferManager:
    """数据传输管理器"""

    def __init__(self, main_dir: str = ".", transfer_interval: int = 600, ui_callback=None):
        """
        初始化数据传输管理器

        Args:
            main_dir: 主目录路径
            transfer_interval: 传输间隔（秒），默认600秒=10分钟
            ui_callback: UI回调函数，用于刷新界面显示
        """
        self.main_dir = Path(main_dir)
        self.transfer_interval = transfer_interval
        self.scripts_dir = self.main_dir / "generated_scripts"
        
        # 主目录文件
        self.main_excel_file = self.main_dir / "主数据表.xlsx"
        self.main_clicked_file = self.main_dir / "pdd_automation" / "data" / "main_image_hashes.json"
        self.main_keywords_file = self.main_dir / "已搜索关键词.json"
        self.transfer_queue_dir = self.main_dir / "transfer_queue"
        
        # 确保目录存在
        self.main_dir.mkdir(exist_ok=True)
        self.transfer_queue_dir.mkdir(exist_ok=True)

        # 🔥 修复：初始化主文件（如果不存在）
        self._init_main_files()

        # 传输状态
        self.is_running = False
        self.transfer_thread = None

        # 🔥 UI回调函数
        self.ui_callback = ui_callback

        print(f"数据传输管理器初始化完成")
        print(f"   主目录: {self.main_dir}")
        print(f"   传输间隔: {self.transfer_interval}秒")
        print(f"   UI回调: {'已设置' if ui_callback else '未设置'}")

        # 🔥 强制确保关键方法存在
        self._ensure_critical_methods()

    def _ensure_critical_methods(self):
        """🔥 确保关键方法存在，防止动态删除或导入问题"""
        try:
            # 检查并修复_refresh_ui_searched_keywords方法
            if not hasattr(self, '_refresh_ui_searched_keywords') or not callable(getattr(self, '_refresh_ui_searched_keywords', None)):
                print("⚠️ 检测到_refresh_ui_searched_keywords方法缺失，强制修复...")

                def _refresh_ui_searched_keywords_fixed(self):
                    """🔥 修复版本的刷新UI方法"""
                    try:
                        if not self.ui_callback:
                            print("ℹ️ 未设置UI回调，跳过界面刷新")
                            return

                        # 读取主已搜索关键词文件
                        if not self.main_keywords_file.exists():
                            print("ℹ️ 主关键词文件不存在，跳过界面刷新")
                            return

                        import json
                        with open(self.main_keywords_file, 'r', encoding='utf-8') as f:
                            keywords_data = json.load(f)

                        searched_keywords = set(keywords_data.get('searched_keywords', []))

                        if not searched_keywords:
                            print("ℹ️ 没有已搜索关键词，跳过界面刷新")
                            return

                        # 🔥 调用UI回调函数刷新界面
                        print(f"刷新UI显示: {len(searched_keywords)} 个已搜索关键词")
                        self.ui_callback(searched_keywords)

                    except Exception as e:
                        print(f"❌ 刷新UI显示失败: {e}")

                # 强制绑定方法
                import types
                self._refresh_ui_searched_keywords = types.MethodType(_refresh_ui_searched_keywords_fixed, self)
                print("✅ 已强制修复_refresh_ui_searched_keywords方法")

            print(f"方法检查完成，_refresh_ui_searched_keywords存在: {hasattr(self, '_refresh_ui_searched_keywords')}")

        except Exception as e:
            print(f"❌ 关键方法检查失败: {e}")

    def _init_main_files(self):
        """🔥 初始化主文件（如果不存在）"""
        try:
            # 初始化已点击商品文件
            if not self.main_clicked_file.exists():
                # 确保data目录存在
                self.main_clicked_file.parent.mkdir(exist_ok=True)
                
                initial_clicked_data = {
                    'last_updated': datetime.now().isoformat(),
                    'hashes': [],
                    'total_count': 0,
                    'created': datetime.now().isoformat()
                }
                with open(self.main_clicked_file, 'w', encoding='utf-8') as f:
                    json.dump(initial_clicked_data, f, ensure_ascii=False, indent=2)
                print(f"✅ 初始化主已点击商品文件: {self.main_clicked_file}")

            # 初始化已搜索关键词文件
            if not self.main_keywords_file.exists():
                initial_keywords_data = {
                    'last_updated': datetime.now().isoformat(),
                    'searched_keywords': [],
                    'total_count': 0,
                    'created': datetime.now().isoformat()
                }
                with open(self.main_keywords_file, 'w', encoding='utf-8') as f:
                    json.dump(initial_keywords_data, f, ensure_ascii=False, indent=2)
                print(f"✅ 初始化主已搜索关键词文件: {self.main_keywords_file}")

        except Exception as e:
            print(f"❌ 初始化主文件失败: {e}")

    def start_auto_transfer(self):
        """启动自动传输 - 改为启动时传输一次就停止"""
        if self.is_running:
            print("⚠️ 自动传输已在运行中")
            return
        
        self.is_running = True
        self.transfer_thread = threading.Thread(target=self._transfer_once, daemon=True)
        self.transfer_thread.start()
        print(f"启动时传输已启动，将执行一次数据传输后自动停止")
    
    def stop_auto_transfer(self):
        """停止自动传输"""
        self.is_running = False
        if self.transfer_thread:
            self.transfer_thread.join(timeout=5)
        print("自动传输已停止")
    
    def _transfer_once(self):
        """🔥 优化：启动时传输一次就停止，减少内存使用"""
        try:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] 开始启动时数据传输...")
            
            # 🔥 内存优化：分批处理，避免一次性加载所有数据
            import gc
            
            # 1. 收集所有浏览器数据（优化版）
            transfer_data = self._collect_browser_data_optimized()
            
            # 2. 处理传输数据
            if transfer_data and transfer_data.get('browsers'):
                print(f"发现 {len(transfer_data['browsers'])} 个浏览器的数据")
                
                # 🔥 逐个处理浏览器数据，处理完一个立即释放内存
                for browser_id, browser_data in transfer_data['browsers'].items():
                    try:
                        self._process_single_transfer({
                            'browser_id': browser_id,
                            'new_clicked': browser_data.get('new_clicked', []),
                            'new_keywords': browser_data.get('new_keywords', []),
                            'new_products': browser_data.get('new_products', [])
                        })
                        
                        # 🔥 处理完一个浏览器后立即清理内存
                        browser_data.clear()
                        gc.collect()
                        
                    except Exception as browser_e:
                        print(f"❌ 处理浏览器 {browser_id} 数据失败: {browser_e}")
                        continue
                
                # 🔥 清理传输数据
                transfer_data.clear()
                gc.collect()
                
                print(f"✅ 启动时数据传输完成")
            else:
                print("ℹ️ 没有发现需要传输的数据")
            
        except Exception as e:
            print(f"❌ 启动时数据传输异常: {e}")
        finally:
            # 传输完成后自动停止
            self.is_running = False
            print("启动时传输已完成，自动停止")
            
            # 🔥 最终内存清理
            import gc
            collected = gc.collect()
            print(f"传输完成后清理了 {collected} 个对象")
    
    def manual_transfer(self):
        """手动触发传输（按需传输）"""
        if self.is_running:
            print("⚠️ 传输正在进行中，请等待完成")
            return False
        
        print("手动触发数据传输...")
        self.start_auto_transfer()
        return True
    
    def _process_pending_transfers(self):
        """处理待处理的传输任务"""
        try:
            if not self.transfer_queue_dir.exists():
                return
                
            # 查找传输队列中的任务
            transfer_files = list(self.transfer_queue_dir.glob("transfer_*.json"))
            
            for transfer_file in transfer_files:
                try:
                    print(f"处理传输任务: {transfer_file.name}")
                    
                    # 读取传输数据
                    with open(transfer_file, 'r', encoding='utf-8') as f:
                        transfer_data = json.load(f)
                    
                    # 处理传输任务
                    self._process_single_transfer(transfer_data)
                    
                    # 删除已处理的传输文件
                    transfer_file.unlink()
                    print(f"✅ 传输任务完成: {transfer_file.name}")
                    
                except Exception as e:
                    print(f"❌ 处理传输任务失败 {transfer_file.name}: {e}")
                    
        except Exception as e:
            print(f"❌ 处理待处理传输失败: {e}")
    
    def _process_single_transfer(self, transfer_data: Dict):
        """处理单个传输任务"""
        try:
            browser_id = transfer_data.get('browser_id', 'unknown')
            print(f"处理浏览器 {browser_id} 的传输数据")
            
            # 1. 处理已点击商品（增量式保存）
            if 'new_clicked' in transfer_data and transfer_data['new_clicked']:
                self._incremental_save_clicked_products(browser_id, transfer_data['new_clicked'])
            
            # 2. 处理已搜索关键词（增量式保存）
            if 'new_keywords' in transfer_data and transfer_data['new_keywords']:
                self._incremental_save_searched_keywords(browser_id, transfer_data['new_keywords'])
            
            # 3. 处理新商品数据（增量式保存）
            if 'new_products' in transfer_data and transfer_data['new_products']:
                self._incremental_save_new_products(browser_id, transfer_data['new_products'])
            
            print(f"✅ 浏览器 {browser_id} 传输数据处理完成")
            
        except Exception as e:
            print(f"❌ 处理传输数据失败: {e}")
    
    def _incremental_save_clicked_products(self, browser_id: str, new_clicked: List[str]):
        """增量式保存已点击商品"""
        try:
            # 读取现有数据
            existing_data = {}
            if self.main_clicked_file.exists():
                with open(self.main_clicked_file, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            
            # 获取现有哈希值
            existing_hashes = set(existing_data.get('hashes', []))
            
            # 添加新的哈希值
            new_hashes = set(new_clicked)
            combined_hashes = existing_hashes | new_hashes
            
            # 计算新增数量
            added_count = len(combined_hashes - existing_hashes)
            
            # 保存增量数据
            updated_data = {
                'last_updated': datetime.now().isoformat(),
                'hashes': sorted(list(combined_hashes)),
                'total_count': len(combined_hashes),
                'new_added': added_count,
                'last_browser_update': browser_id
            }
            
            with open(self.main_clicked_file, 'w', encoding='utf-8') as f:
                json.dump(updated_data, f, ensure_ascii=False, indent=2)
            
            print(f"✅ 已点击商品增量保存完成: 新增 {added_count} 个，总计 {len(combined_hashes)} 个")
            
        except Exception as e:
            print(f"❌ 增量保存已点击商品失败: {e}")
    
    def _incremental_save_searched_keywords(self, browser_id: str, new_keywords: List[str]):
        """增量式保存已搜索关键词"""
        try:
            # 读取现有数据
            existing_data = {}
            if self.main_keywords_file.exists():
                with open(self.main_keywords_file, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            
            # 获取现有关键词
            existing_keywords = set(existing_data.get('searched_keywords', []))
            
            # 添加新的关键词
            new_keywords_set = set(new_keywords)
            combined_keywords = existing_keywords | new_keywords_set
            
            # 计算新增数量
            added_count = len(combined_keywords - existing_keywords)
            
            # 保存增量数据
            updated_data = {
                'last_updated': datetime.now().isoformat(),
                'searched_keywords': sorted(list(combined_keywords)),
                'total_count': len(combined_keywords),
                'new_added': added_count,
                'last_browser_update': browser_id
            }
            
            with open(self.main_keywords_file, 'w', encoding='utf-8') as f:
                json.dump(updated_data, f, ensure_ascii=False, indent=2)
            
            print(f"✅ 已搜索关键词增量保存完成: 新增 {added_count} 个，总计 {len(combined_keywords)} 个")
            
            # 🔥 传输完成后，清理浏览器端的已搜索关键词
            self._clean_browser_searched_keywords(browser_id, new_keywords)
            
            # 刷新UI显示
            if self.ui_callback:
                self.ui_callback(combined_keywords)
            
        except Exception as e:
            print(f"❌ 增量保存已搜索关键词失败: {e}")
    
    def _clean_browser_searched_keywords(self, browser_id: str, searched_keywords: List[str]):
        """🔥 清理浏览器端的已搜索关键词（传输后删除）"""
        try:
            # 找到浏览器目录
            browser_dir = self.scripts_dir / f"browser_{browser_id}"
            if not browser_dir.exists():
                print(f"⚠️ 浏览器目录不存在: {browser_dir}")
                return
            
            # 读取浏览器的配置文件
            config_file = browser_dir / "config_api.json"
            if not config_file.exists():
                print(f"⚠️ 浏览器配置文件不存在: {config_file}")
                return
            
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # 获取当前关键词列表
            current_keywords = config.get('parse_settings', {}).get('search_keywords', [])
            if not current_keywords:
                print(f"ℹ️ 浏览器 {browser_id} 没有搜索关键词")
                return
            
            # 清理已搜索的关键词
            cleaned_keywords = []
            removed_count = 0
            
            for keyword in current_keywords:
                # 如果关键词以"---已搜索"结尾，则删除
                if keyword.endswith('---已搜索'):
                    removed_count += 1
                    print(f"删除已搜索关键词: {keyword}")
                else:
                    cleaned_keywords.append(keyword)
            
            # 更新配置文件
            if 'parse_settings' not in config:
                config['parse_settings'] = {}
            
            config['parse_settings']['search_keywords'] = cleaned_keywords
            
            # 保存更新后的配置
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            print(f"✅ 浏览器 {browser_id} 已搜索关键词清理完成: 删除了 {removed_count} 个关键词")
            
        except Exception as e:
            print(f"❌ 清理浏览器已搜索关键词失败: {e}")
            import traceback
            print(f"详细错误: {traceback.format_exc()}")
    
    def _incremental_save_new_products(self, browser_id: str, new_products: List[Dict]):
        """增量式保存新商品数据"""
        try:
            # 这里可以添加Excel增量保存逻辑
            # 暂时只记录日志
            print(f"新商品数据: {len(new_products)} 个（来自浏览器 {browser_id}）")
            
        except Exception as e:
            print(f"❌ 增量保存新商品数据失败: {e}")
    
    def receive_browser_data(self, browser_id: str, data: Dict) -> bool:
        """🔥 接收浏览器主动传输的数据"""
        try:
            print(f"接收浏览器 {browser_id} 的传输数据")
            
            # 创建传输任务文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            task_filename = f"transfer_{timestamp}_{browser_id}.json"
            task_file = self.transfer_queue_dir / task_filename
            
            # 添加浏览器ID和时间戳
            transfer_data = {
                'browser_id': browser_id,
                'timestamp': datetime.now().isoformat(),
                'new_clicked': data.get('hashes', []),
                'new_keywords': data.get('searched_keywords', []),
                'new_products': data.get('new_products', [])
            }
            
            # 保存传输任务
            with open(task_file, 'w', encoding='utf-8') as f:
                json.dump(transfer_data, f, ensure_ascii=False, indent=2)
            
            print(f"✅ 传输任务已创建: {task_filename}")
            return True
            
        except Exception as e:
            print(f"❌ 接收浏览器数据失败: {e}")
            return False
    
    def get_transfer_status(self) -> Dict:
        """🔥 获取传输状态信息"""
        try:
            status = {
                'is_running': self.is_running,
                'transfer_interval': self.transfer_interval,
                'last_check': getattr(self, '_last_check_time', '未开始'),
                'queue_count': 0,
                'main_files': {}
            }
            
            # 统计传输队列中的任务数量
            if self.transfer_queue_dir.exists():
                transfer_files = list(self.transfer_queue_dir.glob("transfer_*.json"))
                status['queue_count'] = len(transfer_files)
            
            # 检查主文件状态
            if self.main_clicked_file.exists():
                with open(self.main_clicked_file, 'r', encoding='utf-8') as f:
                    clicked_data = json.load(f)
                    status['main_files']['clicked_products'] = {
                        'total_count': clicked_data.get('total_count', 0),
                        'last_updated': clicked_data.get('last_updated', '未知')
                    }
            
            if self.main_keywords_file.exists():
                with open(self.main_keywords_file, 'r', encoding='utf-8') as f:
                    keywords_data = json.load(f)
                    status['main_files']['searched_keywords'] = {
                        'total_count': keywords_data.get('total_count', 0),
                        'last_updated': keywords_data.get('last_updated', '未知')
                    }
            
            return status
            
        except Exception as e:
            print(f"❌ 获取传输状态失败: {e}")
            return {'error': str(e)}
    
    def _collect_browser_data(self) -> Dict:
        """收集所有浏览器的数据"""
        transfer_data = {
            'timestamp': datetime.now().isoformat(),
            'browsers': {},
            'total_new_products': 0,
            'total_new_clicked': 0,
            'total_new_keywords': 0
        }
        
        if not self.scripts_dir.exists():
            return transfer_data
        
        for browser_dir in self.scripts_dir.iterdir():
            if browser_dir.is_dir() and browser_dir.name.startswith('browser_'):
                browser_id = browser_dir.name.replace('browser_', '')
                browser_data = self._collect_single_browser_data(browser_dir, browser_id)
                
                if browser_data and isinstance(browser_data, dict):
                    transfer_data['browsers'][browser_id] = browser_data
                    transfer_data['total_new_products'] += len(browser_data.get('new_products', []))
                    transfer_data['total_new_clicked'] += len(browser_data.get('new_clicked', []))
                    transfer_data['total_new_keywords'] += len(browser_data.get('new_keywords', []))
                elif browser_data is not None:
                    print(f"⚠️ 浏览器 {browser_id} 数据格式错误: {type(browser_data)}")
        
        return transfer_data
    
    def _collect_browser_data_optimized(self) -> Dict:
        """🔥 优化版：内存友好的浏览器数据收集"""
        transfer_data = {
            'timestamp': datetime.now().isoformat(),
            'browsers': {},
            'total_new_products': 0,
            'total_new_clicked': 0,
            'total_new_keywords': 0
        }
        
        if not self.scripts_dir.exists():
            return transfer_data
        
        import gc
        
        for browser_dir in self.scripts_dir.iterdir():
            if browser_dir.is_dir() and browser_dir.name.startswith('browser_'):
                browser_id = browser_dir.name.replace('browser_', '')
                
                # 🔥 单独处理每个浏览器，避免内存积累
                browser_data = self._collect_single_browser_data_optimized(browser_dir, browser_id)
                
                if browser_data and isinstance(browser_data, dict):
                    if (browser_data.get('new_products') or 
                        browser_data.get('new_clicked') or 
                        browser_data.get('new_keywords')):
                        
                        transfer_data['browsers'][browser_id] = browser_data
                        transfer_data['total_new_products'] += len(browser_data.get('new_products', []))
                        transfer_data['total_new_clicked'] += len(browser_data.get('new_clicked', []))
                        transfer_data['total_new_keywords'] += len(browser_data.get('new_keywords', []))
                    
                    # 🔥 立即清理不需要的中间数据
                    gc.collect()
                elif browser_data is not None:
                    print(f"⚠️ 浏览器 {browser_id} 数据格式错误: {type(browser_data)}")
        
        return transfer_data
    
    def _collect_single_browser_data_optimized(self, browser_dir: Path, browser_id: str) -> Optional[Dict]:
        """🔥 优化版：内存友好的单个浏览器数据收集"""
        try:
            browser_data = {
                'browser_id': browser_id,
                'browser_dir': str(browser_dir),
                'new_products': [],
                'new_clicked': [],
                'new_keywords': []
            }
            
            # 1. 🔥 轻量级Excel数据收集（只收集必要信息）
            excel_files = list(browser_dir.glob("output/cj*.xlsx"))
            if excel_files:
                # 只处理最新的文件，避免内存过载
                latest_excel = max(excel_files, key=lambda f: f.stat().st_mtime)
                new_products = self._extract_new_products_from_excel_lightweight(latest_excel)
                browser_data['new_products'] = new_products[:100]  # 限制数量，避免内存过载
            
            # 2. 🔥 轻量级已点击商品收集
            clicked_file = browser_dir / "data" / "main_image_hashes.json"
            if clicked_file.exists() and clicked_file.stat().st_size < 10 * 1024 * 1024:  # 限制文件大小10MB
                new_clicked = self._extract_new_clicked_products(clicked_file)
                browser_data['new_clicked'] = new_clicked[:1000]  # 限制数量
            
            # 3. 🔥 轻量级已搜索关键词收集
            config_file = browser_dir / "config_api.json"
            if config_file.exists() and config_file.stat().st_size < 1024 * 1024:  # 限制文件大小1MB
                new_keywords = self._extract_new_searched_keywords(config_file)
                browser_data['new_keywords'] = new_keywords[:50]  # 限制数量
            
            # 只返回有新数据的浏览器
            if (browser_data['new_products'] or 
                browser_data['new_clicked'] or 
                browser_data['new_keywords']):
                return browser_data
            
            return None
            
        except Exception as e:
            print(f"❌ 收集浏览器 {browser_id} 数据失败: {e}")
            return None
    
    def _extract_new_products_from_excel_lightweight(self, excel_file: Path) -> List[Dict]:
        """🔥 轻量级Excel数据提取，减少内存使用"""
        try:
            # 检查文件大小，避免处理过大的文件
            file_size = excel_file.stat().st_size
            if file_size > 50 * 1024 * 1024:  # 50MB限制
                print(f"⚠️ Excel文件过大，跳过处理: {excel_file}")
                return []
            
            if HAS_PANDAS:
                # 使用pandas分块读取，减少内存使用
                try:
                    df = pd.read_excel(excel_file, chunksize=100)  # 分块读取
                    existing_ids = self._get_existing_product_ids()
                    new_products = []
                    
                    for chunk in df:
                        for _, row in chunk.iterrows():
                            if len(new_products) >= 50:  # 限制处理数量
                                break
                            product_id = str(row.get('商品ID', ''))
                            if product_id and product_id not in existing_ids:
                                new_products.append(row.to_dict())
                        if len(new_products) >= 50:
                            break
                    
                    return new_products
                except:
                    # 如果分块读取失败，使用原始方法但限制行数
                    df = pd.read_excel(excel_file, nrows=200)  # 只读前200行
                    existing_ids = self._get_existing_product_ids()
                    new_products = []
                    for _, row in df.iterrows():
                        product_id = str(row.get('商品ID', ''))
                        if product_id and product_id not in existing_ids:
                            new_products.append(row.to_dict())
                    return new_products
            else:
                # 原有逻辑保持不变
                return self._extract_new_products_from_excel(excel_file)
                
        except Exception as e:
            print(f"❌ 轻量级提取Excel新商品失败: {e}")
            return []
    
    def _collect_single_browser_data(self, browser_dir: Path, browser_id: str) -> Optional[Dict]:
        """收集单个浏览器的数据"""
        try:
            browser_data = {
                'browser_id': browser_id,
                'browser_dir': str(browser_dir),
                'new_products': [],
                'new_clicked': [],
                'new_keywords': []
            }
            
            # 1. 收集EXCEL数据（新增商品）
            excel_files = list(browser_dir.glob("output/cj*.xlsx"))
            for excel_file in excel_files:
                if excel_file.exists():
                    new_products = self._extract_new_products_from_excel(excel_file)
                    browser_data['new_products'].extend(new_products)
            
            # 2. 收集已点击商品
            clicked_file = browser_dir / "data" / "main_image_hashes.json"
            if clicked_file.exists():
                new_clicked = self._extract_new_clicked_products(clicked_file)
                browser_data['new_clicked'] = new_clicked
            
            # 3. 收集已搜索关键词
            config_file = browser_dir / "config_api.json"
            if config_file.exists():
                new_keywords = self._extract_new_searched_keywords(config_file)
                browser_data['new_keywords'] = new_keywords
            
            # 只返回有新数据的浏览器
            if (browser_data['new_products'] or 
                browser_data['new_clicked'] or 
                browser_data['new_keywords']):
                return browser_data
            
            return None
            
        except Exception as e:
            print(f"❌ 收集浏览器 {browser_id} 数据失败: {e}")
            return None
    
    def _extract_new_products_from_excel(self, excel_file: Path) -> List[Dict]:
        """从EXCEL文件提取新商品数据"""
        try:
            if HAS_PANDAS:
                # 使用pandas读取
                df = pd.read_excel(excel_file)
                existing_ids = self._get_existing_product_ids()
                new_products = []
                for _, row in df.iterrows():
                    product_id = str(row.get('商品ID', ''))
                    if product_id and product_id not in existing_ids:
                        new_products.append(row.to_dict())
                return new_products
            elif HAS_OPENPYXL:
                # 使用openpyxl读取
                from openpyxl import load_workbook
                wb = load_workbook(excel_file, read_only=True)
                ws = wb.active
                
                # 获取列标题
                headers = [cell.value for cell in ws[1]]
                product_id_col = None
                for i, header in enumerate(headers):
                    if header == '商品ID':
                        product_id_col = i + 1
                        break
                
                if product_id_col is None:
                    print("❌ 未找到商品ID列")
                    return []
                
                existing_ids = self._get_existing_product_ids()
                new_products = []
                
                for row in ws.iter_rows(min_row=2):
                    product_id = str(row[product_id_col - 1].value) if row[product_id_col - 1].value else ''
                    if product_id and product_id not in existing_ids:
                        product_data = {}
                        for i, cell in enumerate(row):
                            if i < len(headers):
                                product_data[headers[i]] = cell.value
                        new_products.append(product_data)
                
                wb.close()
                return new_products
            else:
                print("❌ 缺少pandas和openpyxl，无法读取Excel文件")
                return []
            
        except Exception as e:
            print(f"❌ 提取EXCEL新商品失败: {e}")
            return []
    
    def _extract_new_clicked_products(self, clicked_file: Path) -> List[str]:
        """提取新的已点击商品ID"""
        try:
            with open(clicked_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 🔥 检查data的类型，确保是字典
            if not isinstance(data, dict):
                print(f"⚠️ main_image_hashes.json文件格式错误，期望字典但得到: {type(data)}")
                print(f"⚠️ 文件内容: {data[:100] if isinstance(data, list) else str(data)[:100]}...")
                return []
            
            browser_clicked = set(data.get('hashes', []))
            main_clicked = self._get_existing_clicked_ids()
            
            # 返回新的已点击商品ID
            new_clicked = list(browser_clicked - main_clicked)
            return new_clicked
            
        except Exception as e:
            print(f"❌ 提取新点击商品失败: {e}")
            import traceback
            print(f"详细错误: {traceback.format_exc()}")
            return []
    
    def _extract_new_searched_keywords(self, config_file: Path) -> List[str]:
        """提取新的已搜索关键词"""
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # 🔥 检查config的类型，确保是字典
            if not isinstance(config, dict):
                print(f"⚠️ config_api.json文件格式错误，期望字典但得到: {type(config)}")
                return []
            
            keywords = config.get('parse_settings', {}).get('search_keywords', [])
            searched_keywords = [kw.replace('---已搜索', '') for kw in keywords if kw.endswith('---已搜索')]
            
            # 获取主文件中已有的关键词
            existing_keywords = self._get_existing_searched_keywords()
            
            # 返回新的已搜索关键词
            new_keywords = [kw for kw in searched_keywords if kw not in existing_keywords]
            return new_keywords
            
        except Exception as e:
            print(f"❌ 提取新搜索关键词失败: {e}")
            import traceback
            print(f"详细错误: {traceback.format_exc()}")
            return []
    
    def _get_existing_product_ids(self) -> Set[str]:
        """获取主数据表中已有的商品ID"""
        try:
            if not self.main_excel_file.exists():
                return set()
                
            if HAS_PANDAS:
                # 使用pandas读取
                df = pd.read_excel(self.main_excel_file)
                return set(str(id) for id in df.get('商品ID', []) if pd.notna(id))
            elif HAS_OPENPYXL:
                # 使用openpyxl读取
                from openpyxl import load_workbook
                wb = load_workbook(self.main_excel_file, read_only=True)
                ws = wb.active
                
                # 获取列标题
                headers = [cell.value for cell in ws[1]]
                product_id_col = None
                for i, header in enumerate(headers):
                    if header == '商品ID':
                        product_id_col = i + 1
                        break
                
                if product_id_col is None:
                    wb.close()
                    return set()
                
                existing_ids = set()
                for row in ws.iter_rows(min_row=2):
                    product_id = row[product_id_col - 1].value
                    if product_id:
                        existing_ids.add(str(product_id))
                
                wb.close()
                return existing_ids
            else:
                print("❌ 缺少pandas和openpyxl，无法读取Excel文件")
                return set()
                
        except Exception as e:
            print(f"❌ 获取已有商品ID失败: {e}")
            return set()
    
    def _get_existing_clicked_ids(self) -> Set[str]:
        """获取主文件中已有的已点击商品ID"""
        try:
            if self.main_clicked_file.exists():
                with open(self.main_clicked_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                return set(data.get('hashes', []))
            return set()
        except Exception as e:
            print(f"❌ 获取已有点击ID失败: {e}")
            return set()
    
    def _get_existing_searched_keywords(self) -> Set[str]:
        """获取主文件中已有的已搜索关键词"""
        try:
            if self.main_keywords_file.exists():
                with open(self.main_keywords_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                return set(data.get('searched_keywords', []))
            return set()
        except Exception as e:
            print(f"❌ 获取已有搜索关键词失败: {e}")
            return set()
    
    def _create_transfer_task(self, transfer_data: Dict) -> str:
        """创建传输任务"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        task_id = f"transfer_{timestamp}"
        
        task_file = self.transfer_queue_dir / f"{task_id}.json"
        with open(task_file, 'w', encoding='utf-8') as f:
            json.dump(transfer_data, f, ensure_ascii=False, indent=2)
        
        return task_id

    def _process_transfer_task(self, task_id: str, transfer_data: Dict):
        """处理传输任务"""
        try:
            print(f"处理传输任务: {task_id}")

            # 1. 更新主EXCEL文件（增量式）
            if transfer_data['total_new_products'] > 0:
                self._update_main_excel(transfer_data)

            # 2. 更新已点击商品文件
            if transfer_data['total_new_clicked'] > 0:
                self._update_main_clicked(transfer_data)

            # 3. 更新已搜索关键词文件
            if transfer_data['total_new_keywords'] > 0:
                self._update_main_keywords(transfer_data)

            # 4. 回传已点击商品到各浏览器（覆盖机制）
            self._distribute_clicked_products()

            # 5. 清理传输任务
            task_file = self.transfer_queue_dir / f"{task_id}.json"
            if task_file.exists():
                task_file.unlink()

            # 6. 🔥 刷新UI显示已搜索关键词
            try:
                # 🔥 动态修复：如果方法不存在，动态添加
                if not hasattr(self, '_refresh_ui_searched_keywords'):
                    print(f"⚠️ 检测到方法缺失，动态修复...")

                    # 动态添加缺失的方法
                    def _refresh_ui_searched_keywords_dynamic(self):
                        """🔥 动态添加的刷新UI方法"""
                        try:
                            if not self.ui_callback:
                                print("ℹ️ 未设置UI回调，跳过界面刷新")
                                return

                            # 读取主已搜索关键词文件
                            if not self.main_keywords_file.exists():
                                print("ℹ️ 主关键词文件不存在，跳过界面刷新")
                                return

                            import json
                            with open(self.main_keywords_file, 'r', encoding='utf-8') as f:
                                keywords_data = json.load(f)

                            searched_keywords = set(keywords_data.get('searched_keywords', []))

                            if not searched_keywords:
                                print("ℹ️ 没有已搜索关键词，跳过界面刷新")
                                return

                            # 🔥 调用UI回调函数刷新界面
                            print(f"刷新UI显示: {len(searched_keywords)} 个已搜索关键词")
                            self.ui_callback(searched_keywords)

                        except Exception as e:
                            print(f"❌ 刷新UI显示失败: {e}")

                    # 绑定方法到实例
                    import types
                    self._refresh_ui_searched_keywords = types.MethodType(_refresh_ui_searched_keywords_dynamic, self)
                    print(f"✅ 已动态修复_refresh_ui_searched_keywords方法")

                self._refresh_ui_searched_keywords()
            except Exception as refresh_e:
                print(f"❌ 刷新UI失败: {refresh_e}")
                import traceback
                print(f"❌ 详细错误: {traceback.format_exc()}")

            print(f"✅ 传输任务完成: {task_id}")

        except Exception as e:
            print(f"❌ 处理传输任务失败: {e}")
            import traceback
            print(f"❌ 详细错误信息: {traceback.format_exc()}")

    def _update_main_excel(self, transfer_data: Dict):
        """更新主EXCEL文件（增量式保存）"""
        try:
            print(f"更新主EXCEL文件...")

            # 收集所有新商品数据
            all_new_products = []
            for browser_data in transfer_data['browsers'].values():
                all_new_products.extend(browser_data['new_products'])

            if not all_new_products:
                return

            # 使用文件锁确保安全
            with self._file_lock(self.main_excel_file):
                if HAS_PANDAS:
                    # 使用pandas处理
                    if self.main_excel_file.exists():
                        existing_df = pd.read_excel(self.main_excel_file)
                    else:
                        existing_df = pd.DataFrame()

                    new_df = pd.DataFrame(all_new_products)
                    
                    if not existing_df.empty:
                        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                    else:
                        combined_df = new_df

                    if '商品ID' in combined_df.columns:
                        combined_df = combined_df.drop_duplicates(subset=['商品ID'], keep='last')

                    combined_df.to_excel(self.main_excel_file, index=False, engine='openpyxl')
                    
                elif HAS_OPENPYXL:
                    # 使用openpyxl处理
                    from openpyxl import Workbook, load_workbook
                    
                    if self.main_excel_file.exists():
                        # 读取现有数据
                        wb = load_workbook(self.main_excel_file)
                        ws = wb.active
                        
                        # 获取现有数据
                        existing_data = []
                        headers = [cell.value for cell in ws[1]]
                        
                        for row in ws.iter_rows(min_row=2):
                            row_data = {}
                            for i, cell in enumerate(row):
                                if i < len(headers):
                                    row_data[headers[i]] = cell.value
                            existing_data.append(row_data)
                        
                        wb.close()
                    else:
                        existing_data = []
                        headers = list(all_new_products[0].keys()) if all_new_products else []
                    
                    # 合并数据
                    combined_data = existing_data + all_new_products
                    
                    # 去重（基于商品ID）
                    if '商品ID' in headers:
                        seen_ids = set()
                        unique_data = []
                        for item in combined_data:
                            product_id = str(item.get('商品ID', ''))
                            if product_id and product_id not in seen_ids:
                                seen_ids.add(product_id)
                                unique_data.append(item)
                        combined_data = unique_data
                    
                    # 保存数据
                    wb = Workbook()
                    ws = wb.active
                    
                    # 写入标题行
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    # 写入数据行
                    for row_idx, row_data in enumerate(combined_data, 2):
                        for col_idx, header in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                    
                    wb.save(self.main_excel_file)
                    wb.close()
                    
                else:
                    print("❌ 缺少pandas和openpyxl，无法保存Excel文件")
                    return

                print(f"✅ 主EXCEL文件更新完成: 新增 {len(all_new_products)} 条数据")

        except Exception as e:
            print(f"❌ 更新主EXCEL文件失败: {e}")

    def _update_main_clicked(self, transfer_data: Dict):
        """更新主已点击商品文件"""
        try:
            print(f"更新已点击商品文件...")

            # 收集所有新点击的商品ID
            all_new_clicked = []
            for browser_data in transfer_data['browsers'].values():
                all_new_clicked.extend(browser_data['new_clicked'])

            if not all_new_clicked:
                return

            # 使用文件锁确保安全
            with self._file_lock(self.main_clicked_file):
                # 读取现有数据
                if self.main_clicked_file.exists():
                    with open(self.main_clicked_file, 'r', encoding='utf-8') as f:
                        existing_data = json.load(f)
                else:
                    existing_data = {
                        'last_updated': datetime.now().isoformat(),
                        'hashes': [],
                        'total_count': 0
                    }

                # 合并新数据
                existing_clicked = set(existing_data.get('hashes', []))
                new_clicked = set(all_new_clicked)
                combined_clicked = existing_clicked | new_clicked

                # 更新数据
                updated_data = {
                    'last_updated': datetime.now().isoformat(),
                    'hashes': sorted(list(combined_clicked)),
                    'total_count': len(combined_clicked),
                    'new_added': len(new_clicked)
                }

                # 保存文件
                with open(self.main_clicked_file, 'w', encoding='utf-8') as f:
                    json.dump(updated_data, f, ensure_ascii=False, indent=2)

                print(f"✅ 已点击商品文件更新完成: 新增 {len(new_clicked)} 个ID")

        except Exception as e:
            print(f"❌ 更新已点击商品文件失败: {e}")

    def _update_main_keywords(self, transfer_data: Dict):
        """更新主已搜索关键词文件"""
        try:
            print(f"更新已搜索关键词文件...")

            # 收集所有新搜索的关键词
            all_new_keywords = []
            for browser_data in transfer_data['browsers'].values():
                all_new_keywords.extend(browser_data['new_keywords'])

            if not all_new_keywords:
                return

            # 使用文件锁确保安全
            with self._file_lock(self.main_keywords_file):
                # 读取现有数据
                if self.main_keywords_file.exists():
                    with open(self.main_keywords_file, 'r', encoding='utf-8') as f:
                        existing_data = json.load(f)
                else:
                    existing_data = {
                        'last_updated': datetime.now().isoformat(),
                        'searched_keywords': [],
                        'total_count': 0
                    }

                # 合并新数据
                existing_keywords = set(existing_data.get('searched_keywords', []))
                new_keywords = set(all_new_keywords)
                combined_keywords = existing_keywords | new_keywords

                # 更新数据
                updated_data = {
                    'last_updated': datetime.now().isoformat(),
                    'searched_keywords': sorted(list(combined_keywords)),
                    'total_count': len(combined_keywords),
                    'new_added': len(new_keywords)
                }

                # 保存文件
                with open(self.main_keywords_file, 'w', encoding='utf-8') as f:
                    json.dump(updated_data, f, ensure_ascii=False, indent=2)

                print(f"✅ 已搜索关键词文件更新完成: 新增 {len(new_keywords)} 个关键词")
                
                # 🔥 更新完成后立即刷新UI
                self._refresh_ui_searched_keywords()

        except Exception as e:
            print(f"❌ 更新已搜索关键词文件失败: {e}")

    def _distribute_clicked_products(self):
        """分发已点击商品到各浏览器（回传覆盖机制）"""
        try:
            print(f"开始分发已点击商品到各浏览器...")

            # 读取主已点击商品文件
            if not self.main_clicked_file.exists():
                print("ℹ️ 主已点击商品文件不存在，跳过分发")
                return

            with open(self.main_clicked_file, 'r', encoding='utf-8') as f:
                main_clicked_data = json.load(f)

            main_clicked_products = main_clicked_data.get('hashes', [])

            if not main_clicked_products:
                print("ℹ️ 没有已点击商品需要分发")
                return

            # 分发到各浏览器目录
            distributed_count = 0
            for browser_dir in self.scripts_dir.iterdir():
                if browser_dir.is_dir() and browser_dir.name.startswith('browser_'):
                    browser_id = browser_dir.name.replace('browser_', '')

                    # 创建浏览器的已点击商品文件
                    browser_clicked_file = browser_dir / "data" / "main_image_hashes.json"
                    browser_clicked_file.parent.mkdir(exist_ok=True)

                    # 准备浏览器的已点击数据
                    browser_clicked_data = {
                        'browser_id': browser_id,
                        'last_updated': datetime.now().isoformat(),
                        'hashes': main_clicked_products,
                        'total_count': len(main_clicked_products),
                        'source': 'main_distribution'
                    }

                    # 覆盖保存到浏览器目录（加文件锁避免并发覆盖）
                    with self._file_lock(browser_clicked_file):
                        with open(browser_clicked_file, 'w', encoding='utf-8') as f:
                            json.dump(browser_clicked_data, f, ensure_ascii=False, indent=2)

                    distributed_count += 1

            print(f"✅ 已点击商品分发完成: 分发到 {distributed_count} 个浏览器")

        except Exception as e:
            print(f"❌ 分发已点击商品失败: {e}")

    def _file_lock(self, file_path: Path):
        """文件锁上下文管理器"""
        return FileLock(file_path)


class FileLock:
    """跨平台文件锁"""

    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.lock_file = file_path.with_suffix(file_path.suffix + '.lock')
        self.lock_handle = None

    def __enter__(self):
        try:
            self.lock_handle = open(self.lock_file, 'w')

            # 🔥 跨平台文件锁处理
            if platform.system() == 'Windows' and HAS_MSVCRT:
                # Windows文件锁
                msvcrt.locking(self.lock_handle.fileno(), msvcrt.LK_NBLCK, 1)
            elif HAS_FCNTL:
                # Linux/Mac文件锁
                fcntl.flock(self.lock_handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            else:
                # 无文件锁支持，使用简单的文件存在检查
                print("⚠️ 文件锁不可用，使用简单锁机制")

            return self

        except (IOError, OSError) as e:
            if self.lock_handle:
                self.lock_handle.close()
            raise Exception(f"获取文件锁失败: {e}")

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.lock_handle:
            try:
                if platform.system() == 'Windows' and HAS_MSVCRT:
                    msvcrt.locking(self.lock_handle.fileno(), msvcrt.LK_UNLCK, 1)
                elif HAS_FCNTL:
                    fcntl.flock(self.lock_handle.fileno(), fcntl.LOCK_UN)
                # 无文件锁时不需要特殊处理
            except:
                pass
            finally:
                self.lock_handle.close()
                if self.lock_file.exists():
                    try:
                        self.lock_file.unlink()
                    except:
                        pass

    def _refresh_ui_searched_keywords(self):
        """🔥 刷新UI中的已搜索关键词显示"""
        try:
            if not self.ui_callback:
                print("ℹ️ 未设置UI回调，跳过界面刷新")
                return

            # 读取主已搜索关键词文件
            if not self.main_keywords_file.exists():
                print("ℹ️ 主关键词文件不存在，跳过界面刷新")
                return

            with open(self.main_keywords_file, 'r', encoding='utf-8') as f:
                keywords_data = json.load(f)

            searched_keywords = set(keywords_data.get('searched_keywords', []))

            if not searched_keywords:
                print("ℹ️ 没有已搜索关键词，跳过界面刷新")
                return

            # 🔥 调用UI回调函数刷新界面
            print(f"刷新UI显示: {len(searched_keywords)} 个已搜索关键词")
            self.ui_callback(searched_keywords)

        except Exception as e:
            print(f"❌ 刷新UI显示失败: {e}")

    def set_ui_callback(self, callback):
        """设置UI回调函数"""
        self.ui_callback = callback
        print("✅ UI回调函数已设置")
